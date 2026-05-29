from __future__ import annotations
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from pydantic import BaseModel
from sqlalchemy.orm import Session

from database import get_db
from models import (
    Campaign, SubdomainScore, TransformationItem,
    CriterionResponse, Domain, Subdomain
)
from scoring import compute_all_scores, compute_domain_score, score_to_bucket
from generation import generate_transformation_items
from pdf_export import generate_roadmap_pdf, generate_sheets_pdf, generate_gantt_pdf, generate_full_report_pdf as generate_rapport_complet_pdf

router = APIRouter(prefix="/campaigns", tags=["synthesis"])


# ─── Scores ───────────────────────────────────────────────────────────────────

@router.get("/{campaign_id}/scores")
def get_scores(campaign_id: str, db: Session = Depends(get_db)):
    campaign = db.get(Campaign, campaign_id)
    if not campaign:
        raise HTTPException(404)

    ss_list = (
        db.query(SubdomainScore)
        .filter_by(campaign_id=campaign_id)
        .all()
    )
    return [
        {
            "subdomain_id":    ss.subdomain_id,
            "subdomain_code":  ss.subdomain.code,
            "subdomain_label": ss.subdomain.label,
            "domain_code":     ss.subdomain.domain.code,
            "score_computed":  round(ss.score_computed, 2) if ss.score_computed else None,
            "score_target":    ss.score_target,
            "gap":             round(ss.score_target - (ss.score_computed or 0), 2),
            "questions_scored":ss.questions_scored,
            "questions_total": ss.questions_total,
            "color_bucket":    score_to_bucket(ss.score_computed),
        }
        for ss in ss_list
    ]


@router.get("/{campaign_id}/scores/radar")
def get_radar(campaign_id: str, db: Session = Depends(get_db)):
    campaign = db.get(Campaign, campaign_id)
    if not campaign:
        raise HTTPException(404)

    domains = (
        db.query(Domain)
        .filter(Domain.code.in_(campaign.domain_scope))
        .order_by(Domain.order_index)
        .all()
    )

    result = []
    for domain in domains:
        domain_score = compute_domain_score(campaign_id, domain.id, db)
        subdomains   = []
        for sd in domain.subdomains:
            ss = (
                db.query(SubdomainScore)
                .filter_by(campaign_id=campaign_id, subdomain_id=sd.id)
                .first()
            )
            score = ss.score_computed if ss else None
            subdomains.append({
                "code":              sd.code,
                "label":             sd.label,
                "score":             round(score, 2) if score else None,
                "target":            ss.score_target if ss else 3.0,
                "gap":               round(ss.score_target - score, 2) if score else None,
                "questions_scored":  ss.questions_scored if ss else 0,
                "questions_total":   ss.questions_total if ss else 0,
                "color_bucket":      score_to_bucket(score),
            })
        result.append({
            "domain_code":  domain.code,
            "domain_label": domain.label,
            "score":        round(domain_score, 2) if domain_score else None,
            "score_pct":    round(domain_score / 4 * 100) if domain_score else None,
            "target":       3.0,
            "subdomains":   subdomains,
        })
    return result


@router.post("/{campaign_id}/scores/recompute")
def recompute_scores(campaign_id: str, db: Session = Depends(get_db)):
    if not db.get(Campaign, campaign_id):
        raise HTTPException(404)
    results = compute_all_scores(campaign_id, db)
    return {"recomputed": len(results), "scores": results}


# ─── Génération chantiers ─────────────────────────────────────────────────────

@router.post("/{campaign_id}/generate")
def generate(campaign_id: str, db: Session = Depends(get_db)):
    campaign = db.get(Campaign, campaign_id)
    if not campaign:
        raise HTTPException(404)
    if campaign.status not in ("IN_PROGRESS", "COMPLETED"):
        raise HTTPException(400, "Campaign must be IN_PROGRESS or COMPLETED")
    return generate_transformation_items(campaign_id, db)


@router.get("/{campaign_id}/transformation-items")
def list_transformation_items(campaign_id: str, db: Session = Depends(get_db)):
    if not db.get(Campaign, campaign_id):
        raise HTTPException(404)
    items = (
        db.query(TransformationItem)
        .filter_by(campaign_id=campaign_id)
        .order_by(TransformationItem.priority_rank)
        .all()
    )
    return [_serialize_item(i) for i in items]


class ItemUpdate(BaseModel):
    effort:           Optional[str] = None
    impact:           Optional[str] = None
    priority_rank:    Optional[int] = None
    status:           Optional[str] = None
    exclusion_reason: Optional[str] = None
    description:      Optional[str] = None


@router.patch("/{campaign_id}/transformation-items/{item_id}")
def update_item(
    campaign_id: str,
    item_id: str,
    body: ItemUpdate,
    db: Session = Depends(get_db),
):
    item = db.get(TransformationItem, item_id)
    if not item or item.campaign_id != campaign_id:
        raise HTTPException(404)

    for field, value in body.model_dump(exclude_none=True).items():
        setattr(item, field, value)
    db.commit()
    return _serialize_item(item)


# ─── Synthèse complète ────────────────────────────────────────────────────────

@router.get("/{campaign_id}/synthesis")
def get_synthesis(campaign_id: str, db: Session = Depends(get_db)):
    campaign = db.get(Campaign, campaign_id)
    if not campaign:
        raise HTTPException(404)

    total    = db.query(CriterionResponse).filter_by(campaign_id=campaign_id).count()
    answered = (
        db.query(CriterionResponse)
        .filter(
            CriterionResponse.campaign_id == campaign_id,
            CriterionResponse.score.isnot(None),
        )
        .count()
    )
    flagged  = (
        db.query(CriterionResponse)
        .filter_by(campaign_id=campaign_id, flagged=True)
        .count()
    )

    items = (
        db.query(TransformationItem)
        .filter_by(campaign_id=campaign_id)
        .order_by(TransformationItem.priority_rank)
        .all()
    )

    # Radar (inline — évite l'import circulaire)
    radar   = _build_radar(campaign_id, campaign.domain_scope, db)
    heatmap = _build_heatmap(campaign_id, campaign.domain_scope, db)

    return {
        "campaign": {
            "id":               campaign.id,
            "title":            campaign.title,
            "status":           campaign.status,
            "consultant_name":  campaign.consultant_name,
            "created_at":       campaign.created_at,
            "synthesis_notes":  campaign.synthesis_notes,
        },
        "supplier": {
            "name":    campaign.supplier.name,
            "sector":  campaign.supplier.sector,
            "country": campaign.supplier.country,
        },
        "stats": {
            "criteria_total":    total,
            "criteria_answered": answered,
            "criteria_flagged":  flagged,
            "completion_pct":    round(answered / total * 100) if total else 0,
        },
        "radar":   radar,
        "heatmap": heatmap,
        "transformation_plan": {
            "proposed": [_serialize_item(i) for i in items if i.status == "proposed"],
            "accepted": [_serialize_item(i) for i in items if i.status == "accepted"],
            "excluded": [_serialize_item(i) for i in items if i.status == "excluded"],
        },
    }


class SynthesisNotesUpdate(BaseModel):
    synthesis_notes: str


@router.patch("/{campaign_id}/synthesis/notes")
def update_synthesis_notes(
    campaign_id: str, body: SynthesisNotesUpdate, db: Session = Depends(get_db)
):
    c = db.get(Campaign, campaign_id)
    if not c:
        raise HTTPException(404)
    c.synthesis_notes = body.synthesis_notes
    db.commit()
    return {"ok": True}


# ─── Export PDF ───────────────────────────────────────────────────────────────

@router.get("/{campaign_id}/report.pdf")
def export_pdf(campaign_id: str, db: Session = Depends(get_db)):
    synthesis = get_synthesis(campaign_id, db)
    try:
        from pdf_export import generate_pdf
        from routers.assessment import get_weak_points
        synthesis["weak_points"] = get_weak_points(campaign_id, db)
        # Pré-calculer les données spider (Jinja2 ne supporte pas les list comprehensions imbriquées)
        synthesis["spider_domain_labels"] = [d["domain_label"][:10] for d in synthesis.get("radar", [])]
        synthesis["spider_domain_scores"] = [d.get("score") or 0 for d in synthesis.get("radar", [])]
        synthesis["spider_sd_labels"]     = [sd["subdomain_code"] for d in synthesis.get("heatmap", []) for sd in d.get("subdomains", [])]
        synthesis["spider_sd_scores"]     = [sd.get("score") or 0 for d in synthesis.get("heatmap", []) for sd in d.get("subdomains", [])]
        pdf_bytes = generate_pdf(synthesis)
        return Response(content=pdf_bytes, media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=assessment_{campaign_id[:8]}.pdf"})
    except Exception as e:
        raise HTTPException(500, f"PDF generation failed: {e}")


@router.get("/{campaign_id}/report-roadmap.pdf")
def export_roadmap_pdf(campaign_id: str, db: Session = Depends(get_db)):
    """Export PDF de la synthèse finale avec matrice effort/impact et chantiers."""
    campaign = db.get(Campaign, campaign_id)
    if not campaign: raise HTTPException(404)

    synthesis = get_synthesis(campaign_id, db)
    from routers.assessment import get_weak_points
    from routers.templates import get_tobe

    tobe = get_tobe(campaign_id, db)

    # Couleurs domaines
    DOMAIN_COLORS = {
        "ORG":  {"bg": "#EEEDFE", "border": "#7F77DD", "text": "#3C3489"},
        "PLAN": {"bg": "#E1F5EE", "border": "#1D9E75", "text": "#085041"},
        "SIM":  {"bg": "#FAEEDA", "border": "#BA7517", "text": "#633806"},
        "IQ":   {"bg": "#FCEBEB", "border": "#E24B4A", "text": "#791F1F"},
        "ME":   {"bg": "#E6F1FB", "border": "#378ADD", "text": "#0C447C"},
    }
    DEFAULT_DC = {"bg": "#F1EFE8", "border": "#888780", "text": "#444441"}

    # Chantiers actifs numérotés
    from models import TransformationItem
    items_raw = db.query(TransformationItem).filter_by(campaign_id=campaign_id)        .order_by(TransformationItem.priority_rank).all()
    active = [i for i in items_raw if i.status != "excluded"]
    numbered = []
    for idx, item in enumerate(active):
        code = (item.domain_codes or [""])[0]
        dc = DOMAIN_COLORS.get(code, DEFAULT_DC)
        numbered.append({
            "id": item.id, "num": idx + 1,
            "label": item.label_custom or item.recommendation_label,
            "description": item.description_custom or item.description or "",
            "effort": item.effort, "impact": item.impact,
            "domain_codes": item.domain_codes,
            "template_impacts": _serialize_template_impacts(item),
            "dc_bg": dc["bg"], "dc_border": dc["border"], "dc_text": dc["text"],
        })

    # Quadrants
    def quadrant(effort, impact):
        eh = effort == "fort"
        ih = impact in ("fort", "moyen")
        if not eh and ih: return "qs"
        if eh and ih:     return "mj"
        if not eh:        return "fi"
        return "dp"

    quadrants = [
        {"id": "qs", "label": "Quick wins",        "sub": "Effort ↓ Impact ↑", "headerBg": "#E1F5EE", "headerColor": "#085041", "items": []},
        {"id": "mj", "label": "Chantiers majeurs", "sub": "Effort ↑ Impact ↑", "headerBg": "#EEEDFE", "headerColor": "#3C3489", "items": []},
        {"id": "fi", "label": "Petits gains",       "sub": "Effort ↓ Impact ↓", "headerBg": "#F1EFE8", "headerColor": "#5F5E5A", "items": []},
        {"id": "dp", "label": "À déprioritiser",      "sub": "Effort ↑ Impact ↓", "headerBg": "#FEF3F2", "headerColor": "#791F1F", "items": []},
    ]
    for item in numbered:
        q = quadrant(item["effort"], item["impact"])
        next(x for x in quadrants if x["id"] == q)["items"].append(item)

    # Domaines groupés
    by_domain = {}
    for item in numbered:
        code = (item["domain_codes"] or ["AUTRE"])[0]
        if code not in by_domain: by_domain[code] = []
        by_domain[code].append(item)

    domain_order = [d["domain_code"] for d in tobe.get("domains", [])]
    domains = []
    for code in domain_order:
        if code not in by_domain: continue
        d_data = next((d for d in tobe.get("domains", []) if d["domain_code"] == code), {})
        dc = DOMAIN_COLORS.get(code, DEFAULT_DC)
        gain = (d_data.get("tobe") or 0) - (d_data.get("as_is") or 0)
        domains.append({
            "code": code, "label": d_data.get("domain_label", code),
            "as_is": d_data.get("as_is"), "tobe": d_data.get("tobe"),
            "gain": gain, "items": by_domain[code],
            "dc_bg": dc["bg"], "dc_border": dc["border"], "dc_text": dc["text"],
        })

    # Heatmap pour tableau
    heatmap_rows = []
    for d in synthesis.get("heatmap", []):
        sds = []
        tobe_domain = next((td for td in tobe.get("domains",[]) if td["domain_code"] == d["domain_code"]), {})
        for sd in d["subdomains"]:
            tobe_sd = next((ts for ts in tobe_domain.get("subdomains",[]) if ts["subdomain_code"] == sd["subdomain_code"]), {})
            gain = (tobe_sd.get("tobe") or 0) - (sd.get("score") or 0)
            sds.append({"label": sd["subdomain_label"], "score": sd["score"],
                        "tobe": tobe_sd.get("tobe"), "gain": gain if gain > 0.05 else None})
        heatmap_rows.append({"label": d["domain_label"], "subdomains": sds})

    # Pré-calculer les données spider pour Jinja2
    tobe_domains = tobe.get("domains", [])
    all_sds = [s for d in tobe_domains for s in (d.get("subdomains") or [])]
    data = {
        "campaign":    synthesis["campaign"],
        "supplier":    synthesis["supplier"],
        "stats":       synthesis["stats"],
        "active_items": numbered,
        "tobe_domains": tobe_domains,
        "quadrants":   quadrants,
        "domains":     domains,
        "heatmap_rows": heatmap_rows,
        # Spider global
        "spider_global_labels": [d.get("domain_label","")[:10] for d in tobe_domains],
        "spider_global_asis":   [d.get("as_is") or 0 for d in tobe_domains],
        "spider_global_tobe":   [d.get("tobe") or d.get("as_is") or 0 for d in tobe_domains],
        # Spider détaillé
        "spider_sd_labels": [s.get("subdomain_code","") for s in all_sds],
        "spider_sd_asis":   [s.get("as_is") or 0 for s in all_sds],
        "spider_sd_tobe":   [s.get("tobe") or s.get("as_is") or 0 for s in all_sds],
    }

    try:
        pdf_bytes = generate_roadmap_pdf(data)
        return Response(content=pdf_bytes, media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=roadmap_{campaign_id[:8]}.pdf"})
    except Exception as e:
        raise HTTPException(500, f"PDF generation failed: {e}")


@router.get("/{campaign_id}/sheets.pdf")
def export_sheets_pdf(campaign_id: str, db: Session = Depends(get_db)):
    """Export PDF des fiches chantiers."""
    campaign = db.get(Campaign, campaign_id)
    if not campaign:
        raise HTTPException(404)

    from routers.sheets import get_campaign_sheets
    sheets_data = get_campaign_sheets(campaign_id, db)

    PHASES = {
        "quick_win": {"label": "Quick wins",  "color": "#1D9E75", "bg": "#E1F5EE"},
        "court":     {"label": "Court terme",  "color": "#7F77DD", "bg": "#EEEDFE"},
        "moyen":     {"label": "Moyen terme", "color": "#BA7517", "bg": "#FAEEDA"},
        "long":      {"label": "Long terme",   "color": "#888780", "bg": "#F1EFE8"},
    }
    DOMAIN_COLORS = {
        "ORG": "#7F77DD", "PLAN": "#1D9E75", "SIM": "#BA7517",
        "IQ": "#E24B4A", "ME": "#378ADD",
    }

    entries = []
    for s in sheets_data:
        phase = s.get("phase", "moyen")
        ph = PHASES.get(phase, PHASES["moyen"])
        dc = (s.get("domain_codes") or [""])[0]
        entries.append({
            **s,
            "phase_label":  ph["label"],
            "phase_color":  ph["color"],
            "phase_bg":     ph["bg"],
            "domain_color": DOMAIN_COLORS.get(dc, "#888780"),
        })

    synthesis = get_synthesis(campaign_id, db)
    data = {
        "campaign": synthesis["campaign"],
        "supplier": synthesis["supplier"],
        "entries":  entries,
    }

    try:
        pdf_bytes = generate_sheets_pdf(data)
        return Response(
            content=pdf_bytes, media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=sheets_{campaign_id[:8]}.pdf"},
        )
    except Exception as e:
        raise HTTPException(500, f"PDF generation failed: {e}")


@router.get("/{campaign_id}/gantt.pdf")
def export_gantt_pdf(campaign_id: str, db: Session = Depends(get_db)):
    """Export PDF du planning Gantt en A4 paysage."""
    campaign = db.get(Campaign, campaign_id)
    if not campaign:
        raise HTTPException(404)

    from routers.gantt import get_gantt
    from datetime import date

    gantt_data = get_gantt(campaign_id, db)
    campaign_data = get_synthesis(campaign_id, db)

    MONTHS_FR = ['Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Jun',
                 'Jul', 'Aoû', 'Sep', 'Oct', 'Nov', 'Déc']

    now = date.today()
    months = []
    y, m = now.year, now.month
    for i in range(18):
        months.append({"year": y, "month": m, "key": f"{y}-{m:02d}", "label": MONTHS_FR[m - 1]})
        m += 1
        if m > 12:
            m = 1
            y += 1

    grid_first = months[0]["key"]

    def month_diff(k1, k2):
        y1, m1 = map(int, k1.split('-'))
        y2, m2 = map(int, k2.split('-'))
        return (y2 - y1) * 12 + (m2 - m1)

    PHASES_META = {
        "quick_win": {"label": "Quick wins",  "color": "#1D9E75", "bg": "#E1F5EE"},
        "court":     {"label": "Court terme",  "color": "#7F77DD", "bg": "#EEEDFE"},
        "moyen":     {"label": "Moyen terme", "color": "#BA7517", "bg": "#FAEEDA"},
        "long":      {"label": "Long terme",   "color": "#888780", "bg": "#F1EFE8"},
    }

    by_phase = []
    for phase_group in gantt_data["by_phase"]:
        ph = PHASES_META.get(phase_group["phase"], PHASES_META["moyen"])
        items = []
        for item in phase_group["items"]:
            if item["start_month"]:
                offset = month_diff(grid_first, item["start_month"])
                dur = item["duration_months"] or 3
                bar_cells = [""] * 18
                for j in range(dur):
                    idx = offset + j
                    if 0 <= idx < 18:
                        bar_cells[idx] = "active"
                items.append({
                    **item,
                    "bar_start":   offset,
                    "bar_width":   dur,
                    "bar_visible": True,
                    "cells":       bar_cells,
                    "phase_color": ph["color"],
                })
            else:
                items.append({
                    **item,
                    "bar_start":   None,
                    "bar_visible": False,
                    "cells":       [""] * 18,
                    "phase_color": ph["color"],
                })
        if items:
            by_phase.append({**ph, "phase": phase_group["phase"], "items": items})

    data = {
        "campaign": campaign_data["campaign"],
        "supplier": campaign_data["supplier"],
        "months":   months,
        "by_phase": by_phase,
    }

    try:
        pdf_bytes = generate_gantt_pdf(data)
        return Response(
            content=pdf_bytes, media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=gantt_{campaign_id[:8]}.pdf"},
        )
    except Exception as e:
        raise HTTPException(500, f"PDF generation failed: {e}")


@router.get("/{campaign_id}/full-report.pdf")
def export_rapport_complet_pdf(campaign_id: str, db: Session = Depends(get_db)):
    """Export PDF complet : planning Gantt (paysage) + fiches chantiers (portrait)."""
    campaign = db.get(Campaign, campaign_id)
    if not campaign:
        raise HTTPException(404)

    from routers.gantt import get_gantt
    from routers.sheets import get_campaign_sheets
    from datetime import date

    # ── Données Gantt ────────────────────────────────────────────────────────
    gantt_raw      = get_gantt(campaign_id, db)
    campaign_data  = get_synthesis(campaign_id, db)

    MONTHS_FR = ['Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Jun',
                 'Jul', 'Aoû', 'Sep', 'Oct', 'Nov', 'Déc']
    now = date.today()
    months, y, m = [], now.year, now.month
    for _ in range(18):
        months.append({"year": y, "month": m, "key": f"{y}-{m:02d}", "label": MONTHS_FR[m - 1]})
        m += 1
        if m > 12:
            m = 1; y += 1

    grid_first = months[0]["key"]

    def month_diff(k1, k2):
        y1, m1 = map(int, k1.split('-'))
        y2, m2 = map(int, k2.split('-'))
        return (y2 - y1) * 12 + (m2 - m1)

    PHASES_META = {
        "quick_win": {"label": "Quick wins",  "color": "#1D9E75", "bg": "#E1F5EE"},
        "court":     {"label": "Court terme",  "color": "#7F77DD", "bg": "#EEEDFE"},
        "moyen":     {"label": "Moyen terme", "color": "#BA7517", "bg": "#FAEEDA"},
        "long":      {"label": "Long terme",   "color": "#888780", "bg": "#F1EFE8"},
    }

    by_phase = []
    for pg in gantt_raw["by_phase"]:
        ph = PHASES_META.get(pg["phase"], PHASES_META["moyen"])
        items = []
        for item in pg["items"]:
            if item["start_month"]:
                offset = month_diff(grid_first, item["start_month"])
                dur = item["duration_months"] or 3
                cells = [""] * 18
                for j in range(dur):
                    idx = offset + j
                    if 0 <= idx < 18:
                        cells[idx] = "active"
                items.append({**item, "bar_visible": True, "cells": cells, "phase_color": ph["color"]})
            else:
                items.append({**item, "bar_visible": False, "cells": [""] * 18, "phase_color": ph["color"]})
        if items:
            by_phase.append({**ph, "phase": pg["phase"], "items": items})

    gantt_data = {
        "campaign": campaign_data["campaign"],
        "supplier": campaign_data["supplier"],
        "months":   months,
        "by_phase": by_phase,
    }

    # ── Données fiches ───────────────────────────────────────────────────────
    PHASES_SHEETS = {
        "quick_win": {"label": "Quick wins",  "color": "#1D9E75"},
        "court":     {"label": "Court terme",  "color": "#7F77DD"},
        "moyen":     {"label": "Moyen terme", "color": "#BA7517"},
        "long":      {"label": "Long terme",   "color": "#888780"},
    }
    DOMAIN_COLORS = {
        "ORG": "#7F77DD", "PLAN": "#1D9E75", "SIM": "#BA7517",
        "IQ": "#E24B4A", "ME": "#378ADD",
    }

    sheets_raw = get_campaign_sheets(campaign_id, db)
    entries = []
    for s in sheets_raw:
        phase = s.get("phase", "moyen")
        ph    = PHASES_SHEETS.get(phase, PHASES_SHEETS["moyen"])
        dc    = (s.get("domain_codes") or [""])[0]
        entries.append({
            **s,
            "phase_label":  ph["label"],
            "phase_color":  ph["color"],
            "domain_color": DOMAIN_COLORS.get(dc, "#888780"),
        })

    sheets_data = {
        "campaign": campaign_data["campaign"],
        "supplier": campaign_data["supplier"],
        "entries":  entries,
    }

    try:
        pdf_bytes = generate_rapport_complet_pdf(gantt_data, sheets_data)
        return Response(
            content=pdf_bytes, media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=rapport_complet_{campaign_id[:8]}.pdf"},
        )
    except Exception as e:
        raise HTTPException(500, f"PDF generation failed: {e}")


@router.get("/{campaign_id}/actions.xlsx")
def export_actions_xlsx(campaign_id: str, db: Session = Depends(get_db)):
    """Export Excel du plan d'actions par domaine / sous-domaine."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import date as date_cls

    campaign = db.get(Campaign, campaign_id)
    if not campaign:
        raise HTTPException(404)

    from routers.sheets import get_campaign_sheets
    sheets_data  = get_campaign_sheets(campaign_id, db)
    camp_summary = get_synthesis(campaign_id, db)

    PHASE_LABELS = {
        "quick_win": "Quick wins", "court": "Court terme",
        "moyen": "Moyen terme",   "long":  "Long terme",
    }
    DOMAIN_HEX = {
        "ORG": "7F77DD", "PLAN": "1D9E75",
        "SIM": "BA7517", "IQ":   "E24B4A", "ME": "378ADD",
    }

    # ── Construire domaine → sous-domaine → chantiers ────────────────────
    items_raw = (
        db.query(TransformationItem)
        .filter_by(campaign_id=campaign_id)
        .filter(TransformationItem.status != "excluded")
        .filter(TransformationItem.template_id.isnot(None))
        .order_by(TransformationItem.priority_rank)
        .all()
    )

    domain_label_cache: dict = {}
    sd_label_cache: dict     = {}

    def get_domain_label(code):
        if code not in domain_label_cache:
            d = db.query(Domain).filter_by(code=code).first()
            domain_label_cache[code] = d.label if d else code
        return domain_label_cache[code]

    def get_sd_label(code):
        if code not in sd_label_cache:
            s = db.query(Subdomain).filter_by(code=code).first()
            sd_label_cache[code] = s.label if s else code
        return sd_label_cache[code]

    domain_order_map = {code: i for i, code in enumerate(campaign.domain_scope)}
    structure: dict  = {}

    for item in items_raw:
        dc = (item.domain_codes    or [""])[0]
        sc = (item.subdomain_codes or [""])[0]
        if dc not in structure:
            structure[dc] = {
                "label": get_domain_label(dc),
                "order": domain_order_map.get(dc, 99),
                "subs":  {},
            }
        if sc not in structure[dc]["subs"]:
            structure[dc]["subs"][sc] = {"label": get_sd_label(sc), "items": []}

        sheet_entry = next((s for s in sheets_data if s["item_id"] == item.id), None)
        actions     = (sheet_entry["sheet"].get("key_actions") or []) if sheet_entry else []

        structure[dc]["subs"][sc]["items"].append({
            "num":    item.priority_rank,
            "label":  item.label_custom or item.recommendation_label,
            "phase":  PHASE_LABELS.get(item.phase or "moyen", "Moyen terme"),
            "effort": item.effort or "—",
            "impact": item.impact or "—",
            "actions": actions,
        })

    sorted_domains = sorted(structure.items(), key=lambda kv: kv[1]["order"])

    # ── Workbook ─────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plan d'actions"

    NCOLS = 8

    def fill(hex6):
        return PatternFill("solid", fgColor=hex6)

    def font(bold=False, color="1A1928", size=10, italic=False):
        return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)

    def align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    _thin = Side(style="thin", color="D0DEEE")

    def border_all():
        return Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    def border_chantier(hex6, is_first):
        left = Side(style="medium", color=hex6) if is_first else _thin
        return Border(left=left, right=_thin, top=_thin, bottom=_thin)

    # Largeurs
    for col, w in zip("ABCDEFGH", [5, 36, 14, 54, 10, 10, 22, 14]):
        ws.column_dimensions[col].width = w

    NAVY  = "12175E"
    BLUE  = "0070AD"
    BLUE2 = "005C96"
    WHITE = "FFFFFF"
    LIGHT = "F0F5FA"
    RESP  = "EEF6FF"

    row = 1

    # Titre
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row, 1, f"Plan d'actions — {camp_summary['supplier']['name']}")
    c.font = font(bold=True, color=WHITE, size=14)
    c.fill = fill(NAVY)
    c.alignment = align("left", "center")
    ws.row_dimensions[row].height = 30
    for col in range(2, NCOLS + 1):
        ws.cell(row, col).fill = fill(NAVY)
    row += 1

    # Méta
    ws.merge_cells(f"A{row}:H{row}")
    parts = [f"Campagne : {camp_summary['campaign']['title']}"]
    if camp_summary["campaign"].get("consultant_name"):
        parts.append(f"Consultant : {camp_summary['campaign']['consultant_name']}")
    parts.append(f"Date : {date_cls.today().strftime('%d/%m/%Y')}")
    c = ws.cell(row, 1, "   " + "     ·     ".join(parts))
    c.font = font(color=WHITE, size=9)
    c.fill = fill(NAVY)
    c.alignment = align("left", "center")
    ws.row_dimensions[row].height = 18
    for col in range(2, NCOLS + 1):
        ws.cell(row, col).fill = fill(NAVY)
    row += 1

    # Ligne vide
    ws.row_dimensions[row].height = 8
    row += 1

    # En-têtes
    headers = ["#", "Chantier", "Phase", "Action", "Effort", "Impact", "Responsable", "Date cible"]
    for ci, h in enumerate(headers, 1):
        bg = BLUE2 if ci >= 7 else BLUE
        c = ws.cell(row, ci, h)
        c.font      = font(bold=True, color=WHITE, size=10)
        c.fill      = fill(bg)
        c.alignment = align("center" if ci in (1, 5, 6) else "left", "center")
        c.border    = border_all()
    ws.row_dimensions[row].height = 22
    ws.auto_filter.ref = f"A{row}:H{row}"
    ws.freeze_panes    = f"A{row + 1}"
    row += 1

    # ── Données ──────────────────────────────────────────────────────────
    action_num = 0

    for dc, domain in sorted_domains:
        domain_hex = DOMAIN_HEX.get(dc, "888780")

        # Header domaine
        ws.merge_cells(f"A{row}:H{row}")
        c = ws.cell(row, 1, f"  {domain['label']}")
        c.font      = font(bold=True, color=WHITE, size=11)
        c.fill      = fill(NAVY)
        c.alignment = align("left", "center")
        ws.row_dimensions[row].height = 24
        for col in range(2, NCOLS + 1):
            ws.cell(row, col).fill = fill(NAVY)
        row += 1

        for sc, sd in domain["subs"].items():
            # Header sous-domaine
            ws.merge_cells(f"A{row}:H{row}")
            c = ws.cell(row, 1, f"      {sd['label']}")
            c.font      = font(bold=True, color=WHITE, size=10)
            c.fill      = fill(BLUE)
            c.alignment = align("left", "center")
            ws.row_dimensions[row].height = 20
            for col in range(2, NCOLS + 1):
                ws.cell(row, col).fill = fill(BLUE)
            row += 1

            for item in sd["items"]:
                actions = item["actions"] or [""]

                for a_idx, action_text in enumerate(actions):
                    action_num += 1
                    is_first = a_idx == 0
                    bg = WHITE if action_num % 2 == 1 else LIGHT

                    # A — numéro
                    c = ws.cell(row, 1, action_num)
                    c.font      = font(color="6B6A7B", size=9)
                    c.fill      = fill(bg)
                    c.alignment = align("center", "center")
                    c.border    = border_all()

                    # B — chantier (label sur la 1re action seulement)
                    lbl = f"[{item['num']}]  {item['label']}" if is_first else ""
                    c = ws.cell(row, 2, lbl)
                    c.font      = font(bold=is_first, size=9)
                    c.fill      = fill(bg)
                    c.alignment = align("left", "center", wrap=True)
                    c.border    = border_chantier(domain_hex, is_first)

                    # C — phase (1re action seulement)
                    c = ws.cell(row, 3, item["phase"] if is_first else "")
                    c.font      = font(size=9, italic=True, color="5F5E5A")
                    c.fill      = fill(bg)
                    c.alignment = align("center", "center")
                    c.border    = border_all()

                    # D — action
                    is_empty  = not action_text
                    disp_text = (action_text if action_text
                                 else ("(no action entered)" if is_first else ""))
                    c = ws.cell(row, 4, disp_text)
                    c.font      = font(size=9, italic=is_empty,
                                       color="9B9AA8" if is_empty else "1A1928")
                    c.fill      = fill(bg)
                    c.alignment = align("left", "center", wrap=True)
                    c.border    = border_all()

                    # E — effort
                    c = ws.cell(row, 5, item["effort"] if is_first else "")
                    c.font      = font(size=9)
                    c.fill      = fill(bg)
                    c.alignment = align("center", "center")
                    c.border    = border_all()

                    # F — impact
                    c = ws.cell(row, 6, item["impact"] if is_first else "")
                    c.font      = font(size=9)
                    c.fill      = fill(bg)
                    c.alignment = align("center", "center")
                    c.border    = border_all()

                    # G — Responsable (vide, fond bleu clair)
                    c = ws.cell(row, 7, "")
                    c.fill   = fill(RESP)
                    c.border = border_all()

                    # H — Date cible (vide, fond bleu clair)
                    c = ws.cell(row, 8, "")
                    c.fill   = fill(RESP)
                    c.border = border_all()

                    ws.row_dimensions[row].height = 18
                    row += 1

        # Séparateur entre domaines
        ws.row_dimensions[row].height = 10
        row += 1

    # ── Export ───────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    slug = camp_summary["supplier"]["name"].replace(" ", "_")[:20]
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=plan_actions_{slug}_{campaign_id[:8]}.xlsx"},
    )

def _serialize_template_impacts(item):
    if not item.template: return []
    return [{"subdomain_code": i.subdomain.code, "maturity_target": i.maturity_target}
            for i in item.template.impacts]


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _serialize_item(item: TransformationItem) -> dict:
    return {
        "id":               item.id,
        "label":            item.recommendation_label,
        "description":      item.description,
        "domain_codes":     item.domain_codes,
        "subdomain_codes":  item.subdomain_codes,
        "effort":           item.effort,
        "impact":           item.impact,
        "priority_rank":    item.priority_rank,
        "status":           item.status,
        "exclusion_reason": item.exclusion_reason,
        "source":           item.source,
    }


def _build_radar(campaign_id: str, domain_scope: list, db: Session) -> list:
    domains = (
        db.query(Domain)
        .filter(Domain.code.in_(domain_scope))
        .order_by(Domain.order_index)
        .all()
    )
    result = []
    for domain in domains:
        domain_score = compute_domain_score(campaign_id, domain.id, db)
        subdomains   = []
        for sd in domain.subdomains:
            ss = (
                db.query(SubdomainScore)
                .filter_by(campaign_id=campaign_id, subdomain_id=sd.id)
                .first()
            )
            score = ss.score_computed if ss else None
            subdomains.append({
                "code":              sd.code,
                "label":             sd.label,
                "score":             round(score, 2) if score else None,
                "target":            ss.score_target if ss else 3.0,
                "gap":               round(ss.score_target - score, 2) if score else None,
                "questions_scored":  ss.questions_scored if ss else 0,
                "questions_total":   ss.questions_total if ss else 0,
                "color_bucket":      score_to_bucket(score),
            })
        result.append({
            "domain_code":  domain.code,
            "domain_label": domain.label,
            "score":        round(domain_score, 2) if domain_score else None,
            "score_pct":    round(domain_score / 4 * 100) if domain_score else None,
            "target":       3.0,
            "subdomains":   subdomains,
        })
    return result


def _build_heatmap(campaign_id: str, domain_scope: list, db: Session) -> list:
    domains = (
        db.query(Domain)
        .filter(Domain.code.in_(domain_scope))
        .order_by(Domain.order_index)
        .all()
    )
    result = []
    for domain in domains:
        sds = []
        for sd in domain.subdomains:
            ss = (
                db.query(SubdomainScore)
                .filter_by(campaign_id=campaign_id, subdomain_id=sd.id)
                .first()
            )
            score = ss.score_computed if ss else None
            sds.append({
                "subdomain_code":   sd.code,
                "subdomain_label":  sd.label,
                "score":            round(score, 2) if score else None,
                "target":           ss.score_target if ss else 3.0,
                "gap":              round(ss.score_target - score, 2) if score else None,
                "questions_scored": ss.questions_scored if ss else 0,
                "questions_total":  ss.questions_total if ss else 0,
                "color_bucket":     score_to_bucket(score),
            })
        result.append({
            "domain_code":  domain.code,
            "domain_label": domain.label,
            "subdomains":   sds,
        })
    return result
